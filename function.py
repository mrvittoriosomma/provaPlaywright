import openpyxl
from playwright.async_api import expect
from playwright.sync_api import Playwright, sync_playwright, expect

def get_data(filename, sheetname):
    workbook = openpyxl.load_workbook(".\\" + filename)
    sheet = workbook[sheetname]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for i in range(2, totalrows + 1):
        dataList = []
        for j in range(1, totalcols + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    return mainList

def loginWordPress(username, password, playwright):
    browser = playwright.chromium.launch(headless=True, args=["--start-maximized"])
    context = browser.new_context(no_viewport=True)
    page = context.new_page()
    page.goto("https://wordpress.com/it/")
    page.screenshot(path="screenshot/screenshot1.png", full_page=True)
    page.get_by_role("menuitem", name="Accedi").click()
    page.screenshot(path="screenshot/screenshot2.png", full_page=True)
    page.get_by_role("button", name="Accetta tutti").click()
    page.screenshot(path="screenshot/screenshot3.png", full_page=True)
    page.get_by_label("Indirizzo e-mail o nome").fill(username)
    page.screenshot(path="screenshot/screenshot4.png", full_page=True)
    page.get_by_role("button", name="Continua", exact=True).click()
    page.screenshot(path="screenshot/screenshot5.png", full_page=True)
    page.get_by_label("Password").fill(password)
    page.screenshot(path="screenshot/screenshot6.png", full_page=True)
    page.get_by_role("button", name="Accedi").click()
    page.screenshot(path="screenshot/screenshot7.png", full_page=True)
    expect(page.get_by_label("Visualizza test32253.")).to_have_text("Titolo del sitotest32253.wordpress.com")
    page.screenshot(path="screenshot/screenshot8.png", full_page=True)
    context.close()
    browser.close()