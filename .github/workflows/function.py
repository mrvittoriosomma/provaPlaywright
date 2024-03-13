import re
import openpyxl

from playwright.async_api import expect
from playwright.sync_api import expect

def get_data():
    workbook = openpyxl.load_workbook(".\\Book1.xlsx")
    sheet = workbook["Sheet1"]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for i in range(2, totalrows + 1):
        dataList = []
        for j in range(1, totalcols + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    return mainList

def loginToDo(username, password, page):
    page.goto("https://app.any.do/")
    page.get_by_placeholder("Enter your work email").fill(username)
    page.locator("form").get_by_role("button").click()
    page.get_by_placeholder("Password").fill(password)
    page.get_by_role("button", name="Sign In").click()

def createTask(task, page):
    page.get_by_test_id("DynamicTextArea - visible textarea").click()
    page.get_by_test_id("DynamicTextArea - visible textarea").fill(task)
    page.locator(".MyDayAddTask__endAdornment > .ButtonInlineText").click()

def checkTask(page):
    page.get_by_label("Check task").first.click()

def delateTask(page):
    page.locator("div").filter(has_text=re.compile(r"^My lists > PersonalProvaUpdate statusList Add task$")).get_by_role("button").click()

def loginWordPress(username, password, page):
    page.goto("https://wordpress.com/it/")
    page.screenshot(path="screenshot/screenshot1.png", full_page=True)
    page.get_by_role("menuitem", name="Accedi").click()
    page.screenshot(path="screenshot/screenshot2.png", full_page=True)
    page.get_by_role("button", name="Accetta tutti").click()
    page.screenshot(path="screenshot/screenshot3.png", full_page=True)
    page.get_by_label("Indirizzo e-mail o nome").fill(username)
    page.screenshot(path="screenshot/screenshot4.png", full_page=True)
    page.get_by_role("button", name="Continua", exact=True).click()
    page.screenshot(path="screenshot/screenshot5.png", full_page=True)
    page.get_by_label("Password").fill(password)
    page.screenshot(path="screenshot/screenshot6.png", full_page=True)
    page.get_by_role("button", name="Accedi").click()
    page.screenshot(path="screenshot/screenshot7.png", full_page=True)
    expect(page.get_by_label("Visualizza test32253.")).to_have_text("Titolo del sitotest32253.wordpress.com")
    page.screenshot(path="screenshot/screenshot8.png", full_page=True)

def richiamami(numeroCellulare, page):
    page.goto("https://eniplenitude.com/")
    page.get_by_title("SCOPRI TREND CASA FREE").get_by_role("button").click()
    page.locator("#carrelloinpagina").get_by_role("button", name="TI CHIAMIAMO NOI").click()
    page.get_by_placeholder("Telefono").fill(numeroCellulare)
    page.get_by_label("Acconsento a ricevere").check()
    page.get_by_role("button", name="RICHIAMAMI").click()
    expect(page.get_by_text("Richiesta inviata, a presto!")).to_have_text(re.compile(r"Richiesta inviata, a presto!"))
    page.get_by_role("button", name="CHIUDI", exact=True).click()

def attivaOfferta(page):
    page.goto("https://eniplenitude.com/")
    page.get_by_label("Accettare le impostazioni di").click()
    page.get_by_title("SCOPRI TREND CASA FREE").get_by_role("button").click()
    page.get_by_role("button", name="ATTIVA TREND CASA FREE").click()
    page.get_by_role("button", name="Conferma e continua").click()
    page.get_by_role("button", name="La fornitura è già attiva con").click()
    page.get_by_role("button", name="Continua", exact=True).click()
    page.get_by_role("button", name="a un'altra persona").click()
    page.get_by_role("button", name="Continua").nth(2).click()
    page.get_by_role("button", name="LA FORNITURA È ATTIVA CON UN ALTRO FORNITORE", exact=True).click()
    page.get_by_role("button", name="Continua").nth(3).click()
    page.get_by_role("button", name="a un'altra persona").nth(1).click()
    page.get_by_role("button", name="Continua").nth(4).click()
    page.get_by_text("Puoi cambiare fornitore solo").click()

def fillForm(page):
    page.goto("https://docs.google.com/forms/d/e/1FAIpQLSeE94CTlX7mIDLNc1b_M46D4utUd2WulcKugx-w9tFxKVJC8A/viewform")
    page.get_by_label("Test 1 *").get_by_label("2").click()
    page.get_by_label("Test 2 *").get_by_label("3").click()
    page.get_by_label("Test 3 *").get_by_label("4").click()
    page.get_by_label("Test 4 *").get_by_label("5").click()
    page.get_by_label("Test 5 *").get_by_label("1").click()
    page.get_by_role("button", name="Invia").click()
    page.get_by_text("La tua risposta è stata").click()
    page.screenshot(path="screenshot.png", full_page=True)
    expect(page.get_by_text("La tua risposta è stata")).to_have_text(re.compile(r"La tua risposta è stata registrata"))

